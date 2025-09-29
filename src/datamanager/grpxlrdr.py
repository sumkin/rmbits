from openpyxl import load_workbook


class GroupExcelReader:


    def __init__(self, fname, origfname):
        self.fname = fname
        self.wb = load_workbook(filename = self.fname)
        self.ws = self.wb['Template']
        self.origfname = origfname


    def read(self):
        rownum = 0
        for row in self.ws.rows:
            if row[0].value is None:
                continue
            if row[0].value.strip() == 'FILL ONLY WHITE CELLS':
                continue
            if row[0].value.strip() == 'POS COUNTRY':
                continue
            pos = row[0].value
            agent = row[1].value
            agent_id = row[2].value
            agent_tier = row[3].value
            product_name = row[4].value
            bc = row[5].value
            customer_type = row[6].value
            pax_cnt = row[7].value

            flights = []
            
            # Outbound. XXX-HEL.
            depdt = row[8].value
            dow = row[9].value
            orgn = row[10].value
            dstn = row[11].value
            fltnum = row[12].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])

            # Outbound. HEL-XXX.
            depdt = row[13].value
            dow = row[14].value
            orgn = row[15].value
            dstn = row[16].value
            fltnum = row[17].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])

            # Inbound. XXX-HEL.
            depdt = row[18].value
            dow = row[19].value
            orgn = row[20].value
            dstn = row[21].value
            fltnum = row[22].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])

            # Inbound. HEL-XXX.
            depdt = row[23].value
            dow = row[24].value
            orgn = row[25].value
            dstn = row[26].value
            fltnum = row[27].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])

            # Segment 5.
            depdt = row[28].value
            dow = row[29].value
            orgn = row[30].value
            dstn = row[31].value
            fltnum = row[32].value
            if fltnum is not None:
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])

            # Segment 6.
            depdt = row[33].value
            dow = row[34].value
            orgn = row[35].value
            dstn = row[36].value
            fltnum = row[37].value
            if fltnum is not None:
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])

            fare_lc = row[38].value
            yr_lc = row[39].value
            currency = row[40].value
            comments = row[41].value
            status = row[42].value

            if pos is not None and len(pos) == 2:
                yield rownum, [pos,agent,agent_id,agent_tier,product_name,bc,customer_type,pax_cnt,flights,fare_lc,yr_lc,currency,comments,status]
            rownum += 1


    def read2(self):
        rownum = 0
        for row in self.ws.rows:
            if row[0].value is None:
                continue
            if row[0].value.strip() == 'FILL ONLY WHITE CELLS':
                continue
            if row[0].value.strip() == 'POS COUNTRY':
                continue
            pos = row[0].value
            agent = row[1].value
            agent_id = row[2].value
            agent_tier = row[3].value
            product_name = row[4].value
            bc = row[5].value
            customer_type = row[6].value
            pax_cnt = row[7].value

            flights = []
            
            # Outbound. XXX-HEL.
            depdt = row[8].value
            dow = row[9].value
            orgn = row[10].value
            dstn = row[11].value
            fltnum = row[12].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])
            else:
                flights.append([None,None,None,None])

            # Outbound. HEL-XXX.
            depdt = row[13].value
            dow = row[14].value
            orgn = row[15].value
            dstn = row[16].value
            fltnum = row[17].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])
            else:
                flights.append([None,None,None,None])

            # Inbound. XXX-HEL.
            depdt = row[18].value
            dow = row[19].value
            orgn = row[20].value
            dstn = row[21].value
            fltnum = row[22].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])
            else:
                flights.append([None,None,None,None])

            # Inbound. HEL-XXX.
            depdt = row[23].value
            dow = row[24].value
            orgn = row[25].value
            dstn = row[26].value
            fltnum = row[27].value
            if fltnum is not None:
                assert depdt is not None
                assert dow is not None
                assert orgn is not None
                assert dstn is not None
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])
            else:
                flights.append([None,None,None,None])

            # Segment 5.
            depdt = row[28].value
            dow = row[29].value
            orgn = row[30].value
            dstn = row[31].value
            fltnum = row[32].value
            if fltnum is not None:
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])
            else:
                flights.append([None,None,None,None])

            # Segment 6.
            depdt = row[33].value
            dow = row[34].value
            orgn = row[35].value
            dstn = row[36].value
            fltnum = row[37].value
            if fltnum is not None:
                fltnum = fltnum[:2] + fltnum[2:].zfill(4)
                flights.append([depdt,orgn,dstn,fltnum])
            else:
                flights.append([None,None,None,None])

            fare_lc = row[38].value
            yr_lc = row[39].value
            currency = row[40].value
            comments = row[41].value
            status = row[42].value

            if pos is not None and len(pos) == 2:
                yield rownum, [pos,agent,agent_id,agent_tier,product_name,bc,customer_type,pax_cnt,flights,fare_lc,yr_lc,currency,comments,status]
            rownum += 1


    def write(self, i, j, val):
        self.ws.cell(row = i, column = j, value = str(val))


    def save(self, fname):
        self.wb.save(fname)


    def get_season(self):
        return str(self.origfname)[:3]


if __name__ == "__main__":
    gxr = GroupExcelReader('/home/ay49514/tmp/gs.xlsx')
    for rownum, r in gxr.read():
        print(rownum, r)
        #gxr.write(rownum + 2, 32, rownum)
        #gxr.write(rownum + 2, 33, rownum * rownum)
        #gxr.write(rownum + 2, 34, rownum * rownum * rownum)
    gxr.save('/home/ay49514/tmp/groups1_mdf.xlsx')


